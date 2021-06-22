import openpyxl


class HandleExcel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]
        cases = []
        data = list(sh.rows)
        # 读表头
        head_data = []
        for item in data[0]:
            head_data.append(item.value)

        # 读内容
        for items in data[1:]:
            body_data = []
            for item in items:
                body_data.append(item.value)
            # 一条用例数据
            case_data = dict(zip(head_data, body_data))
            cases.append(case_data)

        return cases

    def write_excel(self, row, column, value):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)
